#!/usr/bin/env python3
"""
Import eat tracker output into Weight2026_DEXA_calibratedc.xlsx (food tab)

Usage:
  python3 import_food_log.py <clipboard_file.txt> <excel_file.xlsx>

The clipboard file should contain lines in this format (from the "Copy for Log" button):
  2026-03-17|Chicken Breast (6oz)|280|52|1|
  2026-03-17|Greek Yogurt|130|15|2|with honey
  2026-03-17|Whey Protein Drink|130|25|1|post-workout

Format: date|food_name|calories_per|protein_per|qty|notes

Each day's entries are aggregated into a single description string matching the
existing food tab format:  food_name (cal_c, prot_g), food_name (cal_c, prot_g), ...
Total calories and protein are summed and written to the appropriate columns.
"""

import sys
import os
import re
import zipfile
import tempfile
from datetime import datetime
from openpyxl import load_workbook

# Food tab column layout (1-indexed):
#   col 1: row counter
#   col 2: day name (Mon, Tues, etc.)
#   col 3: date
#   col 4: food description
#   col 5: total calories
#   col 6: total protein
#   col 7: burn (not modified)
FOOD_DESC_COL = 4
FOOD_CAL_COL = 5
FOOD_PRO_COL = 6


def parse_clipboard(filepath):
    entries = []
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = line.split('|')
            if len(parts) < 4:
                print(f"  Skipping malformed line: {line}")
                continue
            date_str = parts[0].strip()
            name = parts[1].strip()
            cal = int(parts[2].strip() or 0)
            pro = int(parts[3].strip() or 0)
            qty = int(parts[4].strip() or 1) if len(parts) > 4 and parts[4].strip() else 1
            note = parts[5].strip() if len(parts) > 5 else ""
            entries.append((date_str, name, cal, pro, qty, note))
    return entries


def find_date_row(ws, target_date):
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val is None:
            continue
        if isinstance(cell_val, datetime):
            if cell_val.strftime('%Y-%m-%d') == target_date:
                return row
        elif isinstance(cell_val, str):
            if target_date in cell_val:
                return row
    return None


def build_description(items):
    """Build food description string matching existing format:
    food_name (cal_c, prot_g), food_name (cal_c, prot_g), ...
    """
    parts = []
    for name, cal, pro, qty, note in items:
        total_cal = cal * qty
        total_pro = pro * qty
        prefix = f"{qty} " if qty > 1 else ""
        suffix = f" ({note})" if note else ""
        parts.append(f"{prefix}{name} ({total_cal}c, {total_pro}g){suffix}")
    return ", ".join(parts)


def import_data(clipboard_path, excel_path):
    entries = parse_clipboard(clipboard_path)
    if not entries:
        print("No entries found in clipboard file.")
        return

    print(f"Found {len(entries)} entries to import.")

    # Group entries by date
    by_date = {}
    for date_str, name, cal, pro, qty, note in entries:
        by_date.setdefault(date_str, []).append((name, cal, pro, qty, note))

    wb = load_workbook(excel_path, data_only=True)
    ws = wb['food']

    cell_updates = []  # (row, col_1indexed, value)
    imported = 0
    skipped = 0

    for date_str in sorted(by_date.keys()):
        items = by_date[date_str]
        row = find_date_row(ws, date_str)
        if row is None:
            print(f"  ✗ No row found for date {date_str}")
            skipped += len(items)
            continue

        # Build description, total cal, total protein
        desc = build_description(items)
        total_cal = sum(cal * qty for _, cal, _, qty, _ in items)
        total_pro = sum(pro * qty for _, _, pro, qty, _ in items)

        # Check if there's existing data — append if so
        existing_desc = ws.cell(row=row, column=FOOD_DESC_COL).value
        existing_cal = ws.cell(row=row, column=FOOD_CAL_COL).value
        existing_pro = ws.cell(row=row, column=FOOD_PRO_COL).value

        if existing_desc and str(existing_desc).strip():
            desc = f"{existing_desc}, {desc}"
            total_cal += int(existing_cal or 0)
            total_pro += int(existing_pro or 0)
            print(f"  ⊕ {date_str}: appending {len(items)} items to existing data")
        else:
            day_name = ws.cell(row=row, column=2).value or ""
            print(f"  ✓ {date_str} ({day_name}): {len(items)} items, {total_cal} cal, {total_pro}g protein")

        cell_updates.append((row, FOOD_DESC_COL, desc))
        cell_updates.append((row, FOOD_CAL_COL, str(total_cal)))
        cell_updates.append((row, FOOD_PRO_COL, str(total_pro)))
        imported += len(items)

    wb.close()

    if not cell_updates:
        print("\nNo cell updates to write.")
        return

    print(f"\n{imported} items imported across {len(by_date)} days, {skipped} skipped.")
    print(f"Writing {len(cell_updates)} cell updates...")

    # --- XML surgery on the food sheet (same technique as import_gym_log_final.py) ---
    from lxml import etree as ET
    from openpyxl.utils import get_column_letter

    extract_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(excel_path, 'r') as z:
        z.extractall(extract_dir)

    wb_xml = open(os.path.join(extract_dir, 'xl', 'workbook.xml'), 'r').read()
    rels_xml = open(os.path.join(extract_dir, 'xl', '_rels', 'workbook.xml.rels'), 'r').read()

    rid_match = re.search(r'name="food"[^>]*r:id="(rId\d+)"', wb_xml) or \
                re.search(r'name="food"[^>]*id="(rId\d+)"', wb_xml, re.IGNORECASE)
    if not rid_match:
        print("ERROR: Could not find 'food' sheet in workbook.xml")
        return
    rid = rid_match.group(1)
    target_match = re.search(f'Id="{rid}"[^>]*Target="([^"]+)"', rels_xml) or \
                   re.search(f'Target="([^"]+)"[^>]*Id="{rid}"', rels_xml)
    sheet_rel = target_match.group(1).lstrip('/')
    if sheet_rel.startswith('xl/'): sheet_rel = sheet_rel[3:]

    sheet_path = os.path.join(extract_dir, 'xl', sheet_rel)

    with open(sheet_path, 'rb') as f:
        orig_xml_bytes = f.read()

    tree = ET.parse(sheet_path)
    root = tree.getroot()
    ns = root.nsmap.get(None, 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')

    sheet_data = root.find(f'{{{ns}}}sheetData')
    row_elems = {int(r.get('r')): r for r in sheet_data.findall(f'{{{ns}}}row')}

    def get_row_elem(row_num):
        if row_num not in row_elems:
            row_elem = ET.SubElement(sheet_data, f'{{{ns}}}row')
            row_elem.set('r', str(row_num))
            row_elems[row_num] = row_elem
        return row_elems[row_num]

    def find_cell(row_elem, cell_ref):
        for c in row_elem.findall(f'{{{ns}}}c'):
            if c.get('r') == cell_ref:
                return c
        return None

    updated = 0
    for row_num, col_num, value in cell_updates:
        col_letter = get_column_letter(col_num)
        cell_ref = f"{col_letter}{row_num}"
        val_str = str(value)

        if not val_str:
            continue

        row_elem = get_row_elem(row_num)
        cell_elem = find_cell(row_elem, cell_ref)

        if cell_elem is None:
            cell_elem = ET.SubElement(row_elem, f'{{{ns}}}c')
            cell_elem.set('r', cell_ref)

        for old_v in cell_elem.findall(f'{{{ns}}}v'):
            cell_elem.remove(old_v)
        for old_f in cell_elem.findall(f'{{{ns}}}f'):
            cell_elem.remove(old_f)
        for old_is in cell_elem.findall(f'{{{ns}}}is'):
            cell_elem.remove(old_is)

        # Calories and protein are numeric; description is string
        if col_num in (FOOD_CAL_COL, FOOD_PRO_COL) and val_str.isdigit():
            cell_elem.attrib.pop('t', None)
            v_elem = ET.SubElement(cell_elem, f'{{{ns}}}v')
            v_elem.text = val_str
        else:
            cell_elem.set('t', 'inlineStr')
            is_elem = ET.SubElement(cell_elem, f'{{{ns}}}is')
            t_elem = ET.SubElement(is_elem, f'{{{ns}}}t')
            t_elem.text = val_str

        updated += 1

    # Fix cell ordering within each modified row
    from openpyxl.utils import column_index_from_string
    for row_elem in sheet_data.findall(f'{{{ns}}}row'):
        cell_elems = row_elem.findall(f'{{{ns}}}c')
        if not cell_elems:
            continue

        def _col_sort_key(c):
            ref = c.get('r', '')
            col_letters = ''.join(ch for ch in ref if ch.isalpha())
            return column_index_from_string(col_letters) if col_letters else 0

        sorted_cells = sorted(cell_elems, key=_col_sort_key)
        needs_sort = any(
            _col_sort_key(sorted_cells[i]) != _col_sort_key(cell_elems[i])
            for i in range(len(sorted_cells))
        )
        if needs_sort:
            for c in cell_elems:
                row_elem.remove(c)
            for c in sorted_cells:
                row_elem.append(c)

        if row_elem.get('spans'):
            col_indices = [_col_sort_key(c) for c in sorted_cells]
            if col_indices:
                row_elem.set('spans', f"{min(col_indices)}:{max(col_indices)}")

    # Splice modified sheetData back into original XML
    new_sd_bytes = ET.tostring(sheet_data)
    new_sd_str = new_sd_bytes.decode('utf-8')
    new_sd_str = re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', new_sd_str)

    orig_xml_str = orig_xml_bytes.decode('utf-8')
    patched = re.sub(
        r'<sheetData[^>]*>.*?</sheetData>',
        new_sd_str,
        orig_xml_str,
        flags=re.DOTALL
    )

    with open(sheet_path, 'wb') as f:
        f.write(patched.encode('utf-8'))

    # Rebuild ZIP
    out_path = excel_path
    tmp_zip = excel_path + '.tmp'
    with zipfile.ZipFile(tmp_zip, 'w', zipfile.ZIP_DEFLATED) as zout:
        for dirpath, dirnames, filenames in os.walk(extract_dir):
            for fname in filenames:
                full = os.path.join(dirpath, fname)
                arc = os.path.relpath(full, extract_dir)
                zout.write(full, arc)

    os.replace(tmp_zip, out_path)
    print(f"\n✓ Updated {updated} cells in {out_path}")

    # Cleanup
    import shutil
    shutil.rmtree(extract_dir, ignore_errors=True)


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python3 import_food_log.py <clipboard_file.txt> <excel_file.xlsx>")
        sys.exit(1)
    import_data(sys.argv[1], sys.argv[2])
